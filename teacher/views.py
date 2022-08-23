from django.http import Http404, HttpResponse, HttpResponseNotFound
from django.shortcuts import redirect, render
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from requests import session
from django.contrib.auth.decorators import login_required
from .models import teacher_profile

from school.models import School

import openpyxl
# Create your views here.



# profile of the teacher
# Ayon
@login_required(login_url='/login')
def profile_teacher(request):
    if request.user.is_authenticated:
        # gets the teacher id of the teacher profile and checks if it exists
        if request.GET['teacher_id'] != '' and teacher_profile.objects.filter(id=request.GET['teacher_id']).exists():
            context = {
                'teacher' : teacher_profile.objects.get(id=request.GET['teacher_id'])
            }
            return render(request, 'teacher/profile.html', context)
        else:
            return HttpResponseNotFound()
    else:
        return redirect('/')        


# get all the teachers
# Ayon
@login_required(login_url='/school/login')
def teachers(request):
    if request.user.is_authenticated:
        teachers = teacher_profile.objects.filter(school = School.objects.get(user=User.objects.get(id=request.user.id)))
        return render(request, 'teacher/teachers.html' , {'teachers' : teachers, 'school_id' : request.user.school.id})
    return HttpResponse("Working")


# login the user and set his usertype
# Ayon
def login_teacher(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']    
        user = authenticate(username=username, password=password)
        if user:
            # User is authenticated

            # checks the user is Teacher or not
            if user.groups.filter(name='Teachers').exists():
                # if yes set the user type as Teachers and login

                # log in the teacher
                login(request, user)

                # starts a session teacher and and userType of the session is Teacher
                request.session['teacher'] = {
                'isLoggedIn' : True,
                'userType' : "Teachers"
                }
                return redirect('/teacher?msg=login success')
            else : 
                return redirect('/teacher/login?msg=invalid group')
        else:
            return redirect("/teacher/login?msg=invalid credentials")   
    return render(request, 'teacher/login.html')


# create the teacher account and teacher profile
# Ayon
@login_required(login_url='/school/school-login')
def create_teacher(request):
    # check if the school is logged in or not
    if request.user.is_authenticated and request.GET['id'] != '' and School.objects.filter(id=request.GET['id']).exists():
        if request.method == 'POST':
            first_name = request.POST['first_name']    
            last_name = request.POST['last_name']
            email = request.POST['email']
            username = request.POST['username']
            password = request.POST['password']
            password1 = request.POST['password1']

            if password  != password1:
                return redirect("/teacher/create?msg=password not match")
            else:
                # create a user account for the teacher
                teacher = User.objects.create_user(username, email, password)
                teacher.first_name = first_name
                teacher.last_name = last_name
                teacher.save()   

                school = School.objects.get(id=request.GET['id'])

                # add the teacher to the Teachers group      
                Group.objects.get(name='Teachers').user_set.add(teacher)

                # create the teacher profile for the teacher account
                add_teacher_profile = teacher_profile.objects.create(
                    user = teacher, # the teacher account
                    school = school,
                    img = request.FILES['img'],
                    full_name  = f"{first_name} {last_name}",
                    phone = request.POST['phone'],
                    gender = request.POST['gender'],
                    qualification = request.POST['qualification'],
                    experience = request.POST['experience'],
                    address = request.POST['address'],
                    zipcode = request.POST['zipcode'],
                )

                # add_teacher_profile.school.add(school)

                return redirect(f'/teacher/profile?teacher_id={add_teacher_profile.id}&msg=add success')
        else:
            context = {'school_id' : request.user.school.id}
            return render(request, 'teacher/add-teacher.html', context)
    else:
        return redirect("/school/login")


# update the teacher_profile
# Ayon
@login_required(login_url='/teacher/login')
def edit_teacher(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            user = User.objects.get(id = request.user.id)
            user.first_name = request.POST['first_name'] 
            user.last_name = request.POST['last_name'] 
            user.save()

            if len( request.FILES) != 0 :
                teacher_profile.objects.filter(user=user).update(img=request.FILES['img'])

            teacher = teacher_profile.objects.filter(user=user).update(
                full_name  = request.POST['first_name'] + ' ' + request.POST['last_name'],
                phone = request.POST['phone'],
                gender = request.POST['gender'],
                qualification = request.POST['qualification'],
                experience = request.POST['experience'],
                address = request.POST['address'],
                zipcode = request.POST['zipcode'],
            )
            return redirect('/teacher/update?msg=update success')
        else:
            user = User.objects.get(id = request.user.id)
            teacher = teacher_profile.objects.get(user=user)
            context = {
                'teacher' : teacher,
                'school_id' : teacher.school.id
            }
            return render(request, 'teacher/edit-teacher.html', context)
    return redirect('/')


# logout for the teacher
# Ayon
def logout_teacher(request):
    try :
        del request.session['teacher']
    except:
        pass    
    logout(request)
    return redirect('/')


@login_required(login_url='/school/school-login')
def upload_teacher(request):
    if request.user.is_authenticated and request.GET['id'] != '' and School.objects.filter(id=request.GET['id']).exists():
        if request.method == 'POST':

            if request.FILES['teacher_file']  != '':
                school = School.objects.get(id=request.GET['id'])
                data = openpyxl.load_workbook(request.FILES['teacher_file'])

                data = data.active

                data_list = []

                for row in range(2, data.max_row):
                    for col in data.iter_cols(1, data.max_column):
                        if col[row].value != None:
                            data_list.append(str(col[row].value))
                        else:
                            return redirect('/teacher/add/upload?id=' + str(request.GET['id']))

                    first_name = data_list[0] 
                    last_name =  data_list[1] 
                    email = data_list[2] 
                    username = data_list[9] 
                    password = data_list[10] 


                    full_name  = f"{first_name} {last_name}"
                    phone = int(data_list[3])
                    gender = data_list[4] 
                    qualification = data_list[5] 
                    experience = data_list[6] 
                    address = data_list[7] 
                    zipcode = int(data_list[8])

                    data_list.clear()

                    teacher = User.objects.create_user(username, email, password)
                    teacher.first_name = first_name
                    teacher.last_name = last_name
                    teacher.save()  

                    # add the teacher to the Teachers group      
                    Group.objects.get(name='Teachers').user_set.add(teacher)

                    # create the teacher profile for the teacher account
                    add_teacher_profile = teacher_profile.objects.create(
                        user = teacher, # the teacher account
                        school = school,
                        full_name  = full_name,
                        phone = phone,
                        gender = gender,
                        qualification = qualification,
                        experience = experience,
                        address = address,
                        zipcode = zipcode,
                    ) 
                return redirect("/teacher/list")
            return redirect('/teacher/add/upload?id=' + str(request.GET['id']))
        context = {'school_id' : request.user.school.id}        
        return render(request, 'teacher/add-upload.html', context)    
    return redirect('/school/school-login')    
        

@login_required(login_url='/teacher/login')
def dash(request):
    if request.user.is_authenticated and teacher_profile.objects.filter(user=User.objects.get(id=request.user.id)).exists():
        teacher = teacher_profile.objects.get(id=request.user.teacherprofile.id)
        context = {
            'teacher' : teacher,
            'school_id' : teacher.school.id
        }
        return render(request, 'teacher/index.html', context)
    return redirect('/teacher/login')    