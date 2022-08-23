import datetime
from django.http import Http404, HttpResponse, HttpResponseNotFound
from django.shortcuts import redirect, render
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from requests import session
from django.contrib.auth.decorators import login_required
from .models import teacher_profile

from school.models import School
from student.models import *
import openpyxl
# Create your views here.



# profile of the student
# Ayon
@login_required(login_url='/login')
def profile_student(request):
    if request.user.is_authenticated:
        # gets the teacher id of the teacher profile and checks if it exists
        if request.GET['student_id'] != '' and student_profile.objects.filter(id=request.GET['student_id']).exists():
            context = {
                'student' : student_profile.objects.get(id=request.GET['student_id']),
                'school_id' : request.user.teacherprofile.school.id
            }
            return render(request, 'student/profile.html', context)
        else:
            return HttpResponseNotFound()
    else:
        return redirect('/')        


# get all the student
# Ayon
@login_required(login_url='/school/login')
def students(request):
    if request.user.is_authenticated and request.GET['id'] != '':
        students = student_profile.objects.filter(school = School.objects.get(id=request.GET['id']))
        return render(request, 'student/students.html' , {'student' : students, 'school_id' : request.GET['id']})
    return HttpResponse("Working")


# login the user and set his usertype
# Ayon
def login_student(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']    
        user = authenticate(username=username, password=password)
        if user:
            # User is authenticated

            # checks the user is Studentsor not
            if user.groups.filter(name='Students').exists():
                # if yes set the user type as Students and login

                # log in the teacher
                login(request, user)

                # starts a session teacher and and userType of the session is Teacher
                request.session['student'] = {
                'isLoggedIn' : True,
                'userType' : "Students"
                }
                return redirect('/student?msg=login success')
            else : 
                return redirect('/student/login?msg=invalid group')
        else:
            return redirect("/student/login?msg=invalid credentials")   
    return render(request, 'student/login.html')


# create the student account and student profile
# Ayon
@login_required(login_url='/school/school-login')
def create_student(request):
    # check if the school/teacher is logged in or not
    if request.user.is_authenticated and request.GET['id'] != '' and School.objects.filter(id=request.GET['id']).exists():
        if request.method == 'POST':
            first_name = request.POST['first_name']    
            last_name = request.POST['last_name']
            email = request.POST['email']
            username = request.POST['username']
            password = request.POST['password']
            password1 = request.POST['password1']

            if password  != password1:
                return redirect("/student/add?msg=password not match")
            else:
                # create a user account for the student
                student = User.objects.create_user(username, email, password)
                student.first_name = first_name
                student.last_name = last_name
                student.save()   

                school = School.objects.get(id=request.GET['id'])

                # add the student to the Students group      
                Group.objects.get(name='Students').user_set.add(student)

                # create the student profile for the student account
                add_student_profile = student_profile.objects.create(
                    user = student, # the teacher account
                    school = school,
                    img = request.FILES['img'],
                    full_name  = f"{first_name} {last_name}",
                    phone = request.POST['phone'],
                    gender = request.POST['gender'],

                    dob = request.POST['dob'],
                    standard = request.POST['standard'],
                    roll_no = request.POST['roll_no'],
                    section = request.POST['section'],

                    father_name = request.POST['father_name'],
                    mother_name = request.POST['mother_name'],
                    father_phone = request.POST['father_phone'],
                    mother_phone = request.POST['mother_phone'],
                    address = request.POST['address'],
                    zipcode  = request.POST['zipcode']
                )


                return redirect(f'/student/profile?student_id={add_student_profile.id}&msg=add success')
        else:
            teacher = teacher_profile.objects.get(id=request.user.teacherprofile.id)
            context = {
                'teacher' : teacher,
                'school_id' : teacher.school.id
            }
            return render(request, 'student/add-student.html', context)
    else:
        return redirect("/school/login")


# update the student_profile
# Ayon
@login_required(login_url='/teacher/login')
def edit_student(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            user = User.objects.get(id = request.user.id)
            user.first_name = request.POST['first_name'] 
            user.last_name = request.POST['last_name'] 
            user.save()

            if len(request.FILES) != 0 :
                student_profile.objects.filter(user=user).update(img = request.FILES.get('img'))

            student = student_profile.objects.filter(user=user).update(
                full_name  = f"{request.POST['first_name'] } {request.POST['last_name'] }",
                phone = request.POST['phone'],
                gender = request.POST['gender'],
                dob = request.POST['dob'],
                standard = request.POST['standard'],
                roll_no = request.POST['roll_no'],
                section = request.POST['section'],
                father_name = request.POST['father_name'],
                mother_name = request.POST['mother_name'],
                father_phone = request.POST['father_phone'],
                mother_phone = request.POST['mother_phone'],
                address = request.POST['address']
            )
            return redirect('/student/update?msg=update success')
        else:
            user = User.objects.get(id = request.user.id)
            student = student_profile.objects.get(user=user)
            context = {
                'student' : student,
                'school_id' : student.school.id
            }
            return render(request, 'student/edit-student.html', context)
    return redirect('/')


# logout for the student
# Ayon
def logout_student(request):
    try :
        del request.session['student']
    except:
        pass    
    logout(request)
    return redirect('/')


@login_required(login_url='/teacher/school-login')
def upload_student(request):
    if request.user.is_authenticated and request.GET['id'] != '' and student_profile.objects.filter(id=request.GET['id']).exists():
        if request.method == 'POST':

            if request.FILES['student_file']  != '':
                school = School.objects.get(id=request.GET['id'])
                data = openpyxl.load_workbook(request.FILES['student_file'])

                data = data.active

                data_list = []

                for row in range(2, data.max_row):
                    for col in data.iter_cols(1, data.max_column):
                        if col[row].value != None:
                            data_list.append(str(col[row].value))
                        else:
                            return redirect('/teacher/add/upload?id=' + str(request.GET['id']))

                    first_name = data_list[0] 
                    last_name =  data_list[1] 
                    gender = data_list[2]
                    dob = datetime.datetime.strptime(data_list[3], "%Y-%m-%d %H:%M:%S")
                    dob = datetime.date.strftime(dob, "%Y-%m-%d")
                    standard = data_list[4]
                    phone = int(data_list[5])
                    roll_no = data_list[6]
                    section = data_list[7]

                    father_name = data_list[8]
                    mother_name = data_list[9]

                    father_phone= data_list[10]
                    mother_phone = data_list[11]
                    username = data_list[12] 
                    password = data_list[13] 
                    email = data_list[14] 
                    address = data_list[15] 
                    zipcode = int(data_list[16])

                    full_name  = f"{first_name} {last_name}"

                    data_list.clear()

                    student = User.objects.create_user(username, email, password)
                    student.first_name = first_name
                    student.last_name = last_name
                    student.save()  

                    # add the teacher to the Students group      
                    Group.objects.get(name='Students').user_set.add(student)

                    # create the student profile for the student account
                    add_student_profile = student_profile.objects.create(
                        user = student, # the student account
                        school = school,
                        full_name  = full_name,
                        phone = phone,
                        gender = gender,
                        dob = dob,
                        standard = standard,
                        roll_no = roll_no,
                        section = section,
                        father_name = father_name,
                        mother_name = mother_name,
                        father_phone = father_phone,
                        mother_phone = mother_phone,
                        address = address,
                        zipcode  = zipcode
                    ) 
                return redirect("/student/list?id=" + str(school.id))
            return redirect('/student/add/upload?id=' + str(request.GET['id']))
        context = {'school_id' : request.user.teacherprofile.school.id}        
        return render(request, 'student/add-upload.html', context)    
    return redirect('/school/school-login')    
        

@login_required(login_url='/student/login')
def dash(request):
    if request.user.is_authenticated and student_profile.objects.filter(user=User.objects.get(id=request.user.id)).exists():
        student = student_profile.objects.get(id=request.user.studentprofile.id)
        context = {
            'student' : student,
            'school_id' : student.school.id
        }
        return render(request, 'student/index.html', context)
    return redirect('/student/login')    